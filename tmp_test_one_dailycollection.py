import json, math, base64, urllib.request, urllib.error
from datetime import datetime, date
import openpyxl

excel_path = r'C:\Users\mbugu\Downloads\paul use.xlsx'
endpoint = "http://test.trimline.co.ke:4548/BC240/ODataV4/Company('Rugi')/DailyCollections"
token = base64.b64encode(b'Philip:Password@2030').decode('ascii')
headers = {'Authorization':'Basic '+token,'Accept':'application/json','Content-Type':'application/json'}

wb=openpyxl.load_workbook(excel_path,data_only=True)
ws=wb[wb.sheetnames[0]]
h=[str(c.value).strip() if c.value is not None else '' for c in ws[1]]
idx={k:i for i,k in enumerate(h)}
r=next(ws.iter_rows(min_row=2, values_only=True))

vdate=r[idx['Date']]
if isinstance(vdate, datetime):
    d=vdate.date().isoformat(); dto=vdate.strftime('%Y-%m-%dT%H:%M:%SZ')
elif isinstance(vdate, date):
    d=vdate.isoformat(); dto=d+'T00:00:00Z'
else:
    d=str(vdate)[:10]; dto=d+'T00:00:00Z'
kg=float(r[idx['Net Weight (kg)']])
member=str(r[idx['Member Number']]).strip()
name=str(r[idx['Member Name']]).strip()
crop=str(r[idx['Crop Type']]).strip()

# find starting No
start_no=1
try:
    req=urllib.request.Request(endpoint+'?%24orderby=No%20desc&%24top=1',headers=headers)
    with urllib.request.urlopen(req,timeout=10) as resp:
        data=json.loads(resp.read().decode('utf-8','ignore'))
        if data.get('value'): start_no=int(data['value'][0].get('No',0))+1
except Exception as e:
    print('Could not query start No:',e)

payload={
 'No':start_no,
 'Farmers_Number':member,
 'Farmers_Name':name[:100],
 'Collections_Date':d,
 'Collection_Time':dto,
 'Collection_Number':f'{member}-{d.replace("-","")}-{start_no}'[:30],
 'Coffee_Type':crop[:20],
 'Kg_Collected':kg,
 'Gross':kg,
 'Tare':0,
 'No_of_Bags':int(math.ceil(kg/90.0)),
 'Factory':'',
 'Cancelled':False,
 'Paid':False,
 'Sent':False,
 'Updated':False,
 'ID_Number':'',
 'Delivered_By':name[:50],
 'Collect_Type':'',
 'Crop':crop[:30],
 'Cumm':kg,
 'Can':'',
 'User':'Philip',
 'Comments':'Uploaded from paul use.xlsx'
}
body=json.dumps(payload).encode('utf-8')
req=urllib.request.Request(endpoint,data=body,headers=headers,method='POST')
try:
    with urllib.request.urlopen(req,timeout=15) as resp:
        out=resp.read().decode('utf-8','ignore')
    print('POST OK',resp.status)
    print(out[:500])
except urllib.error.HTTPError as e:
    err=e.read().decode('utf-8','ignore')
    print('HTTP ERROR',e.code)
    print(err[:2000])
except Exception as e:
    print('ERROR',repr(e))
