import argparse
import base64
import datetime as dt
import json
import re
import sys
import time
import urllib.error
import urllib.request

from openpyxl import load_workbook


def normalize_farmer(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"[^A-Za-z0-9]", "", text)
    return text


def normalize_kg(value):
    if value is None:
        return "0"
    try:
        num = float(value)
    except Exception:
        num = 0.0
    text = f"{num:.2f}".replace("-", "N").replace(".", "")
    return re.sub(r"[^A-Za-z0-9]", "", text)


def build_collection_number(farmer_number, kg_value, row_index, used):
    farmer = normalize_farmer(farmer_number)[:12] or "UNK"
    kg = normalize_kg(kg_value)[:8] or "0"
    base = f"C{farmer}K{kg}R{row_index:05d}"
    candidate = base[:30]

    if candidate not in used:
        used.add(candidate)
        return candidate

    for bump in range(1, 100000):
        suffix = f"X{bump:04d}"
        candidate = (base[: 30 - len(suffix)] + suffix)[:30]
        if candidate not in used:
            used.add(candidate)
            return candidate

    raise RuntimeError("Unable to generate unique Collection_Number")


def to_date_iso(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_datetime_offset_iso(value):
    if isinstance(value, dt.datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=dt.timezone.utc)
        else:
            value = value.astimezone(dt.timezone.utc)
        return value.isoformat().replace("+00:00", "Z")
    if isinstance(value, dt.date):
        value = dt.datetime.combine(value, dt.time(0, 0, tzinfo=dt.timezone.utc))
        return value.isoformat().replace("+00:00", "Z")
    return None


def fetch_string_field_set(url, headers, field_name):
    result = set()
    next_url = url
    while next_url:
        req = urllib.request.Request(next_url, headers=headers, method="GET")
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        for row in data.get("value", []):
            value = row.get(field_name)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                result.add(text)

        next_url = data.get("@odata.nextLink")

    return result


def main():
    parser = argparse.ArgumentParser(description="Upload DailyCollections rows from Excel to BC OData.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--url", required=True)
    parser.add_argument("--username", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--factory", default="KARUNDU")
    parser.add_argument("--limit", type=int, default=0, help="Upload only first N data rows (0 = all)")
    parser.add_argument("--max-retries", type=int, default=3)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb.active

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    expected = [
        "Date",
        "Member Number",
        "Member Name",
        "Season",
        "Crop Type",
        "Net Weight (kg)",
    ]
    missing_headers = [h for h in expected if h not in headers]
    if missing_headers:
        print(f"Missing expected headers: {missing_headers}")
        return 2

    idx = {name: headers.index(name) for name in expected}

    token = base64.b64encode(f"{args.username}:{args.password}".encode("utf-8")).decode("ascii")
    base_headers = {
        "Authorization": f"Basic {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    company_base = args.url.rsplit("/", 1)[0]
    farmers_url = f"{company_base}/Farmers?$select=No&$top=20000"
    existing_url = (
        f"{company_base}/DailyCollections?"
        "$select=Collection_Number&$top=20000"
    )

    print("Prefetching Farmers from BC...")
    farmer_numbers = fetch_string_field_set(farmers_url, base_headers, "No")
    print(f"Farmers loaded: {len(farmer_numbers)}")

    print("Prefetching existing DailyCollections keys...")
    existing_collection_numbers = fetch_string_field_set(
        existing_url,
        base_headers,
        "Collection_Number",
    )
    print(f"Existing Collection_Number loaded: {len(existing_collection_numbers)}")

    used_collection_numbers = set()
    ok = 0
    skipped = 0
    skipped_duplicate = 0
    skipped_missing_customer = 0
    failed = 0

    for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if args.limit and (ok + failed + skipped) >= args.limit:
            break

        date_value = row[idx["Date"]]
        member_no = row[idx["Member Number"]]
        member_name = row[idx["Member Name"]]
        season = row[idx["Season"]]
        crop_type = row[idx["Crop Type"]]
        net_weight = row[idx["Net Weight (kg)"]]

        if member_no is None or net_weight is None:
            skipped += 1
            print(f"SKIP row {excel_row_num}: missing Member Number or Net Weight")
            continue

        collection_number = build_collection_number(member_no, net_weight, excel_row_num, used_collection_numbers)

        payload = {
            "Collection_Number": collection_number,
            "Farmers_Number": str(member_no).strip(),
            "Farmers_Name": "" if member_name is None else str(member_name).strip(),
            "Collections_Date": to_date_iso(date_value),
            "Collection_Time": to_datetime_offset_iso(date_value),
            "Coffee_Type": "" if crop_type is None else str(crop_type).strip(),
            "Kg_Collected": float(net_weight),
            "Factory": args.factory,
            "Crop": "" if season is None else str(season).strip(),
            "Collect_Type": "Manual",
        }

        farmer_key = str(member_no).strip()
        if farmer_key not in farmer_numbers:
            skipped_missing_customer += 1
            continue

        if collection_number in existing_collection_numbers:
            skipped_duplicate += 1
            continue

        if args.dry_run:
            print(json.dumps(payload, ensure_ascii=True))
            ok += 1
            continue

        data = json.dumps(payload, ensure_ascii=True).encode("utf-8")

        attempt = 0
        row_done = False
        while attempt <= args.max_retries and not row_done:
            attempt += 1
            req = urllib.request.Request(args.url, data=data, headers=base_headers, method="POST")
            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    _ = resp.read()
                ok += 1
                existing_collection_numbers.add(collection_number)
                row_done = True
                if ok % 100 == 0:
                    print(f"Uploaded {ok} rows...")
            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", errors="ignore")

                if "Internal_EntityWithSameKeyExists" in body:
                    skipped_duplicate += 1
                    existing_collection_numbers.add(collection_number)
                    row_done = True
                    continue

                if (
                    "Internal_InvalidTableRelation" in body
                    and "field Farmers Number" in body
                ):
                    skipped_missing_customer += 1
                    row_done = True
                    continue

                if e.code in (408, 429, 500, 502, 503, 504) and attempt <= args.max_retries:
                    time.sleep(min(2 * attempt, 5))
                    continue

                failed += 1
                print(f"FAIL row {excel_row_num} status={e.code} body={body[:500]}")
                row_done = True
            except Exception as e:
                message = str(e)
                transient = (
                    "WinError 10051" in message
                    or "timed out" in message.lower()
                    or "temporarily unavailable" in message.lower()
                )
                if transient and attempt <= args.max_retries:
                    time.sleep(min(2 * attempt, 5))
                    continue

                failed += 1
                print(f"FAIL row {excel_row_num} error={e}")
                row_done = True

    skipped_total = skipped + skipped_duplicate + skipped_missing_customer
    print(
        "DONE "
        f"ok={ok} skipped_total={skipped_total} "
        f"skipped_missing_fields={skipped} "
        f"skipped_duplicates={skipped_duplicate} "
        f"skipped_missing_customer={skipped_missing_customer} "
        f"failed={failed}"
    )
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
