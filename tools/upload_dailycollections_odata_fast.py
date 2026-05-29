import argparse
import base64
import concurrent.futures
import datetime as dt
import json
import re
import threading
import time
import urllib.error
import urllib.request

from openpyxl import load_workbook


def _is_transient_error(message):
    text = (message or "").lower()
    return (
        "winerror 10051" in text
        or "errno 11001" in text
        or "getaddrinfo failed" in text
        or "name or service not known" in text
        or "timed out" in text
        or "temporarily unavailable" in text
        or "connection reset" in text
        or "unable to connect" in text
    )


def _fetch_page_json(url, headers, max_retries=5):
    for attempt in range(max_retries + 1):
        req = urllib.request.Request(url, headers=headers, method="GET")
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            # Retry transient server-side errors.
            if e.code in (408, 429, 500, 502, 503, 504) and attempt < max_retries:
                time.sleep(min(2 * (attempt + 1), 10))
                continue
            raise
        except Exception as e:
            if _is_transient_error(str(e)) and attempt < max_retries:
                time.sleep(min(2 * (attempt + 1), 10))
                continue
            raise


def normalize_farmer(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"[^A-Za-z0-9]", "", text)
    return text


def normalize_kg(value):
    if value is None:
        return "0"
    try:
        num = float(value)
    except Exception:
        num = 0.0
    text = f"{num:.2f}".replace("-", "N").replace(".", "")
    return re.sub(r"[^A-Za-z0-9]", "", text)


def build_collection_number(farmer_number, kg_value, row_index):
    farmer = normalize_farmer(farmer_number)[:12] or "UNK"
    kg = normalize_kg(kg_value)[:8] or "0"
    return f"C{farmer}K{kg}R{row_index:05d}"[:30]


def to_date_iso(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_datetime_offset_iso(value):
    if isinstance(value, dt.datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=dt.timezone.utc)
        else:
            value = value.astimezone(dt.timezone.utc)
        return value.isoformat().replace("+00:00", "Z")
    if isinstance(value, dt.date):
        value = dt.datetime.combine(value, dt.time(0, 0, tzinfo=dt.timezone.utc))
        return value.isoformat().replace("+00:00", "Z")
    return None


def fetch_string_field_set(url, headers, field_name):
    result = set()
    next_url = url
    while next_url:
        data = _fetch_page_json(next_url, headers)
        for row in data.get("value", []):
            value = row.get(field_name)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                result.add(text)
        next_url = data.get("@odata.nextLink")
    return result


def post_one(url, headers, payload, max_retries):
    data = json.dumps(payload, ensure_ascii=True).encode("utf-8")
    for attempt in range(max_retries + 1):
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                _ = resp.read()
            return "ok", None
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="ignore")
            if "Internal_EntityWithSameKeyExists" in body:
                return "duplicate", None
            if "Internal_InvalidTableRelation" in body and "field Farmers Number" in body:
                return "missing_customer", None
            if e.code in (408, 429, 500, 502, 503, 504) and attempt < max_retries:
                time.sleep(min(2 * (attempt + 1), 5))
                continue
            return "failed", f"HTTP {e.code}: {body[:500]}"
        except Exception as e:
            message = str(e)
            transient = _is_transient_error(message)
            if transient and attempt < max_retries:
                time.sleep(min(2 * (attempt + 1), 5))
                continue
            return "failed", message


def main():
    parser = argparse.ArgumentParser(description="Fast upload DailyCollections from Excel to BC OData.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--url", required=True)
    parser.add_argument("--username", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--factory", default="KARUNDU")
    parser.add_argument("--max-retries", type=int, default=3)
    parser.add_argument("--workers", type=int, default=12)
    args = parser.parse_args()

    token = base64.b64encode(f"{args.username}:{args.password}".encode("utf-8")).decode("ascii")
    headers = {
        "Authorization": f"Basic {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    company_base = args.url.rsplit("/", 1)[0]
    farmers_url = f"{company_base}/Farmers?$select=No&$top=20000"
    existing_url = f"{company_base}/DailyCollections?$select=Collection_Number&$top=20000"

    print("Prefetching Farmers...")
    farmer_numbers = fetch_string_field_set(farmers_url, headers, "No")
    print(f"Farmers loaded: {len(farmer_numbers)}")

    print("Prefetching existing Collection_Number values...")
    existing_collection_numbers = fetch_string_field_set(existing_url, headers, "Collection_Number")
    print(f"Existing keys loaded: {len(existing_collection_numbers)}")

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb.active
    header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    needed = ["Date", "Member Number", "Member Name", "Season", "Crop Type", "Net Weight (kg)"]
    missing = [h for h in needed if h not in header_row]
    if missing:
        print(f"Missing expected headers: {missing}")
        return 2
    idx = {name: header_row.index(name) for name in needed}

    to_upload = []
    skipped_missing_fields = 0
    skipped_missing_customer = 0
    skipped_duplicate = 0

    for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        date_value = row[idx["Date"]]
        member_no = row[idx["Member Number"]]
        member_name = row[idx["Member Name"]]
        season = row[idx["Season"]]
        crop_type = row[idx["Crop Type"]]
        net_weight = row[idx["Net Weight (kg)"]]

        if member_no is None or net_weight is None:
            skipped_missing_fields += 1
            continue

        farmer_key = str(member_no).strip()
        if farmer_key not in farmer_numbers:
            skipped_missing_customer += 1
            continue

        collection_number = build_collection_number(member_no, net_weight, excel_row_num)
        if collection_number in existing_collection_numbers:
            skipped_duplicate += 1
            continue

        payload = {
            "Collection_Number": collection_number,
            "Farmers_Number": farmer_key,
            "Farmers_Name": "" if member_name is None else str(member_name).strip(),
            "Collections_Date": to_date_iso(date_value),
            "Collection_Time": to_datetime_offset_iso(date_value),
            "Coffee_Type": "" if crop_type is None else str(crop_type).strip(),
            "Kg_Collected": float(net_weight),
            "Factory": args.factory,
            "Crop": "" if season is None else str(season).strip(),
            "Collect_Type": "Manual",
        }
        to_upload.append((excel_row_num, collection_number, payload))

    print(f"Prepared rows for upload: {len(to_upload)}")

    ok = 0
    fail = 0
    skipped_runtime_duplicate = 0
    lock = threading.Lock()

    def worker(item):
        row_num, key, payload = item
        status, detail = post_one(args.url, headers, payload, args.max_retries)
        return row_num, key, status, detail

    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        futures = [ex.submit(worker, item) for item in to_upload]
        processed = 0
        for fut in concurrent.futures.as_completed(futures):
            row_num, key, status, detail = fut.result()
            processed += 1
            with lock:
                if status == "ok":
                    ok += 1
                elif status == "duplicate":
                    skipped_runtime_duplicate += 1
                elif status == "missing_customer":
                    skipped_missing_customer += 1
                else:
                    fail += 1
                    print(f"FAIL row {row_num}: {detail}")

            if processed % 200 == 0:
                print(f"Processed {processed}/{len(to_upload)}")

    skipped_total = (
        skipped_missing_fields
        + skipped_missing_customer
        + skipped_duplicate
        + skipped_runtime_duplicate
    )

    print(
        "DONE "
        f"ok={ok} "
        f"skipped_total={skipped_total} "
        f"skipped_missing_fields={skipped_missing_fields} "
        f"skipped_missing_customer={skipped_missing_customer} "
        f"skipped_duplicates={skipped_duplicate + skipped_runtime_duplicate} "
        f"failed={fail}"
    )

    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
