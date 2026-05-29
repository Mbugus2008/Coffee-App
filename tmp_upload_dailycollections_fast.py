import json
import math
import base64
import time
import urllib.request
import urllib.error
from datetime import datetime, date
import openpyxl

excel_path = r'C:\Users\mbugu\Downloads\paul use.xlsx'
endpoint = "http://test.trimline.co.ke:4548/BC240/ODataV4/Company('Rugi')/DailyCollections"

def headers():
    tok = base64.b64encode(b'Philip:Password@2030').decode('ascii')
    return {'Authorization':'Basic '+tok,'Accept':'application/json','Content-Type':'application/json'}

def post(payload):
    body = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(endpoint, method='POST', headers=headers(), data=body)
    with urllib.request.urlopen(req, timeout=6) as r:
        _ = r.read()

def get_start_no():
    try:
        req = urllib.request.Request(endpoint + '?%24orderby=No%20desc&%24top=1', headers=headers())
        with urllib.request.urlopen(req, timeout=6) as r:
            data = json.loads(r.read().decode('utf-8','ignore'))
        rows = data.get('value',[])
        if rows:
            return int(rows[0].get('No',0)) + 1
    except Exception:
        pass
    return int(time.time())

def to_date_time(v):
    if isinstance(v, datetime):
        return v.date().isoformat(), v.strftime('%Y-%m-%dT%H:%M:%SZ')
    if isinstance(v, date):
        d=v.isoformat(); return d, d+'T00:00:00Z'
    s=str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S','%Y-%m-%d','%d/%m/%Y %H:%M','%d/%m/%Y'):
        try:
            dt=datetime.strptime(s,fmt)
            return dt.date().isoformat(), dt.strftime('%Y-%m-%dT%H:%M:%SZ')
        except Exception:
            pass
    return None,None

wb=openpyxl.load_workbook(excel_path,data_only=True)
ws=wb[wb.sheetnames[0]]
cols=[str(c.value).strip() if c.value is not None else '' for c in ws[1]]
idx={h:i for i,h in enumerate(cols)}

no=get_start_no()
created=0
failed=0
skipped=0
errors=[]

for rn,row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    try:
        member=row[idx['Member Number']]
        name=row[idx['Member Name']]
        crop=row[idx['Crop Type']]
        kg=row[idx['Net Weight (kg)']]
        dt=row[idx['Date']]
    except Exception:
        skipped+=1
        continue
    if member is None or kg is None or dt is None:
        skipped+=1
        continue
    try:
        kg=float(kg)
    except Exception:
        skipped+=1
        continue
    d, t = to_date_time(dt)
    if not d or not t:
        skipped+=1
        continue

    m=str(member).strip()
    nm=str(name or '').strip()[:100]
    cp=str(crop or '').strip()[:30]
    payload={
      'No': int(no),
      'Farmers_Number': m,
      'Farmers_Name': nm,
      'Collections_Date': d,
      'Collection_Time': t,
      'Collection_Number': f"{m}-{d.replace('-','')}-{no}"[:30],
      'Coffee_Type': cp[:20],
      'Kg_Collected': kg,
      'Gross': kg,
      'Tare': 0,
      'No_of_Bags': int(math.ceil(kg/90.0)),
      'Factory': '',
      'Cancelled': False,
      'Paid': False,
      'Sent': False,
      'Updated': False,
      'ID_Number': '',
      'Delivered_By': nm[:50],
      'Collect_Type': '',
      'Crop': cp,
      'Cumm': kg,
      'Can': '',
      'User': 'Philip',
      'Comments': 'Uploaded from paul use.xlsx'
    }

    ok=False
    for _ in range(2):
        try:
            post(payload)
            ok=True
            break
        except urllib.error.HTTPError as e:
            if e.code in (400,401,403,404,409,422):
                try:
                    detail=e.read().decode('utf-8','ignore')[:180]
                except Exception:
                    detail=str(e)
                errors.append(f'row {rn} no {no} http {e.code} {detail}')
                break
            time.sleep(0.7)
        except Exception as e:
            last=str(e)
            time.sleep(0.7)
    if ok:
        created+=1
    else:
        failed+=1
    no+=1
    if (created+failed) % 50 == 0:
        print(f'progress attempted={created+failed} created={created} failed={failed} skipped={skipped}')

print(f'FINAL attempted={created+failed} created={created} failed={failed} skipped={skipped}')
if errors:
    print('ERRORS_START')
    for e in errors[:15]:
        print(e)
    print('ERRORS_END')

try:
    req = urllib.request.Request(endpoint + '?%24top=1', headers=headers())
    with urllib.request.urlopen(req, timeout=6) as r:
        check=json.loads(r.read().decode('utf-8','ignore'))
    print('VERIFY_TOP1_OK rows=', len(check.get('value',[])))
except Exception as e:
    print('VERIFY_TOP1_FAIL', str(e)[:160])
