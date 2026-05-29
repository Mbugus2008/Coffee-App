import json, math, base64, urllib.request, urllib.error
from datetime import datetime, date
import openpyxl

excel_path = r'C:\Users\mbugu\Downloads\paul use.xlsx'
endpoint = "http://test.trimline.co.ke:4548/BC240/ODataV4/Company('Rugi')/DailyCollections"

tok = base64.b64encode(b'Philip:Password@2030').decode('ascii')
HEAD = {'Authorization':'Basic '+tok,'Accept':'application/json','Content-Type':'application/json'}

def req_json(url, method='GET', data=None, timeout=8):
    req = urllib.request.Request(url, method=method, headers=HEAD, data=data)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        t = r.read().decode('utf-8','ignore')
    return json.loads(t) if t else {}

def conv_dt(v):
    if isinstance(v, datetime):
        return v.date().isoformat(), v.strftime('%Y-%m-%dT%H:%M:%SZ')
    if isinstance(v, date):
        d=v.isoformat(); return d, d+'T00:00:00Z'
    s=str(v).strip()
    for f in ('%Y-%m-%d %H:%M:%S','%Y-%m-%d','%d/%m/%Y %H:%M','%d/%m/%Y'):
        try:
            d=datetime.strptime(s,f)
            return d.date().isoformat(), d.strftime('%Y-%m-%dT%H:%M:%SZ')
        except Exception:
            pass
    return None,None

start_no = 1
try:
    top = req_json(endpoint + '?%24orderby=No%20desc&%24top=1')
    if top.get('value'):
        start_no = int(top['value'][0]['No']) + 1
except Exception:
    pass

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb[wb.sheetnames[0]]
h=[str(c.value).strip() if c.value is not None else '' for c in ws[1]]
i={k:v for v,k in enumerate(h)}

created=0; failed=0; skipped=0; no=start_no; errs=[]
for rn,row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    member=row[i['Member Number']] if i['Member Number']<len(row) else None
    name=row[i['Member Name']] if i['Member Name']<len(row) else ''
    crop=row[i['Crop Type']] if i['Crop Type']<len(row) else ''
    kg=row[i['Net Weight (kg)']] if i['Net Weight (kg)']<len(row) else None
    dt=row[i['Date']] if i['Date']<len(row) else None
    if member is None or kg is None or dt is None:
        skipped+=1; continue
    try: kg=float(kg)
    except Exception: skipped+=1; continue
    d,t = conv_dt(dt)
    if not d or not t:
        skipped+=1; continue
    m=str(member).strip(); nm=str(name or '').strip()[:100]; cp=str(crop or '').strip()[:30]
    payload={
        'No': int(no),'Farmers_Number': m,'Farmers_Name': nm,'Collections_Date': d,'Collection_Time': t,
        'Collection_Number': f"{m}-{d.replace('-','')}-{no}"[:30],'Coffee_Type': cp[:20],'Kg_Collected': kg,
        'Gross': kg,'Tare': 0,'No_of_Bags': int(math.ceil(kg/90.0)),'Factory': '','Cancelled': False,
        'Paid': False,'Sent': False,'Updated': False,'ID_Number': '','Delivered_By': nm[:50],'Collect_Type': '',
        'Crop': cp,'Cumm': kg,'Can': '','User': 'Philip','Comments': 'Uploaded from paul use.xlsx'
    }
    try:
        req_json(endpoint, method='POST', data=json.dumps(payload).encode('utf-8'))
        created += 1
    except urllib.error.HTTPError as e:
        failed += 1
        if len(errs) < 12:
            try: detail=e.read().decode('utf-8','ignore')[:160]
            except Exception: detail=str(e)
            errs.append(f'row {rn} no {no} http {e.code} {detail}')
    except Exception as e:
        failed += 1
        if len(errs) < 12:
            errs.append(f'row {rn} no {no} err {str(e)[:160]}')
    no += 1
    if (created+failed) % 50 == 0:
        print(f'progress attempted={created+failed} created={created} failed={failed} skipped={skipped}')

print(f'UPLOAD_SUMMARY attempted={created+failed} created={created} failed={failed} skipped={skipped} startNo={start_no}')
if errs:
    print('ERROR_SAMPLES_BEGIN')
    [print(x) for x in errs]
    print('ERROR_SAMPLES_END')

try:
    v=req_json(endpoint + '?%24top=1')
    print('VERIFY_ENDPOINT_OK topRows=', len(v.get('value',[])))
except Exception as e:
    print('VERIFY_ENDPOINT_FAIL', str(e)[:160])
