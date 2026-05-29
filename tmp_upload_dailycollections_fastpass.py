import json, math, base64, urllib.request, urllib.error, time
from datetime import datetime, date
import openpyxl

excel_path = r'C:\Users\mbugu\Downloads\paul use.xlsx'
endpoint = "http://test.trimline.co.ke:4548/BC240/ODataV4/Company('Rugi')/DailyCollections"
token = base64.b64encode(b'Philip:Password@2030').decode('ascii')
headers = {'Authorization':'Basic '+token,'Accept':'application/json','Content-Type':'application/json'}

wb=openpyxl.load_workbook(excel_path,data_only=True)
ws=wb[wb.sheetnames[0]]
h=[str(c.value).strip() if c.value is not None else '' for c in ws[1]]
idx={k:i for i,k in enumerate(h)}
rows=list(ws.iter_rows(min_row=2, values_only=True))
base_no = int(time.time())
created=0
failed=0

for i,r in enumerate(rows, start=1):
    member=r[idx['Member Number']] if idx.get('Member Number') is not None else None
    dt=r[idx['Date']] if idx.get('Date') is not None else None
    kg=r[idx['Net Weight (kg)']] if idx.get('Net Weight (kg)') is not None else None
    name=r[idx['Member Name']] if idx.get('Member Name') is not None else ''
    crop=r[idx['Crop Type']] if idx.get('Crop Type') is not None else ''

    if member is None or dt is None or kg is None:
        print(f'row {i}: skipped missing')
        continue

    try:
        kgf=float(kg)
    except Exception:
        print(f'row {i}: skipped bad kg')
        continue

    if isinstance(dt, datetime):
        d=dt.date().isoformat(); dto=dt.strftime('%Y-%m-%dT%H:%M:%SZ')
    elif isinstance(dt, date):
        d=dt.isoformat(); dto=d+'T00:00:00Z'
    else:
        s=str(dt).strip()
        try:
            p=datetime.strptime(s,'%Y-%m-%d %H:%M:%S')
            d=p.date().isoformat(); dto=p.strftime('%Y-%m-%dT%H:%M:%SZ')
        except Exception:
            print(f'row {i}: skipped bad date')
            continue

    no = base_no + i
    member_s=str(member).strip()
    name_s=str(name or '').strip()
    crop_s=str(crop or '').strip()

    payload={
      'No': no,
      'Farmers_Number': member_s,
      'Farmers_Name': name_s[:100],
      'Collections_Date': d,
      'Collection_Time': dto,
      'Collection_Number': f'{member_s}-{d.replace("-","")}-{no}'[:30],
      'Coffee_Type': crop_s[:20],
      'Kg_Collected': kgf,
      'Gross': kgf,
      'Tare': 0,
      'No_of_Bags': int(math.ceil(kgf/90.0)),
      'Factory': '',
      'Cancelled': False,
      'Paid': False,
      'Sent': False,
      'Updated': False,
      'ID_Number': '',
      'Delivered_By': name_s[:50],
      'Collect_Type': '',
      'Crop': crop_s[:30],
      'Cumm': kgf,
      'Can': '',
      'User': 'Philip',
      'Comments': 'Uploaded from paul use.xlsx'
    }
    body=json.dumps(payload).encode('utf-8')

    try:
        req=urllib.request.Request(endpoint,data=body,headers=headers,method='POST')
        with urllib.request.urlopen(req,timeout=5) as resp:
            _=resp.read()
        created += 1
        print(f'row {i}: created')
    except urllib.error.HTTPError as e:
        msg=e.read().decode('utf-8','ignore')
        if e.code == 400 and ('already exists' in msg or 'EntityWithSameKeyExists' in msg):
            print(f'row {i}: duplicate')
        else:
            failed += 1
            print(f'row {i}: HTTP {e.code}')
    except Exception:
        failed += 1
        print(f'row {i}: network/error')

print(f'FINAL created={created} failed={failed} total_rows={len(rows)}')
