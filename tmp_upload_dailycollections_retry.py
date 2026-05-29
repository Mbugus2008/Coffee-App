import json
import math
import base64
import time
import urllib.request
import urllib.error
from datetime import datetime, date
import openpyxl

excel_path = r'C:\Users\mbugu\Downloads\paul use.xlsx'
endpoint = "http://test.trimline.co.ke:4548/BC240/ODataV4/Company('Rugi')/DailyCollections"
user = 'Philip'
pwd = 'Password@2030'

def auth_header():
    token = base64.b64encode(f"{user}:{pwd}".encode('ascii')).decode('ascii')
    return {
        'Authorization': f'Basic {token}',
        'Accept': 'application/json',
        'Content-Type': 'application/json'
    }

def request_json(url, method='GET', body=None, timeout=20):
    req = urllib.request.Request(url, method=method, headers=auth_header(), data=body)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read().decode('utf-8', 'ignore')
    return json.loads(raw) if raw else {}

def get_start_no():
    try:
        data = request_json(endpoint + '?%24orderby=No%20desc&%24top=1')
        rows = data.get('value', [])
        if rows:
            return int(rows[0].get('No', 0)) + 1
    except Exception:
        pass
    # fallback high range to avoid collisions if lookup fails intermittently
    return int(time.time())

def as_date_and_time(v):
    if isinstance(v, datetime):
        return v.date().isoformat(), v.strftime('%Y-%m-%dT%H:%M:%SZ')
    if isinstance(v, date):
        d = v.isoformat()
        return d, d + 'T00:00:00Z'
    s = str(v).strip()
    if not s:
        return None, None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date().isoformat(), dt.strftime('%Y-%m-%dT%H:%M:%SZ')
        except Exception:
            pass
    return None, None

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}
required = ['Date', 'Member Number', 'Member Name', 'Crop Type', 'Net Weight (kg)']
missing = [h for h in required if h not in idx]
if missing:
    raise RuntimeError('Missing columns: ' + ', '.join(missing))

no = get_start_no()
created = 0
failed = 0
skipped = 0
sample_errors = []

for row_num, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    member_no = r[idx['Member Number']] if idx['Member Number'] < len(r) else None
    member_name = r[idx['Member Name']] if idx['Member Name'] < len(r) else ''
    crop = r[idx['Crop Type']] if idx['Crop Type'] < len(r) else ''
    kg = r[idx['Net Weight (kg)']] if idx['Net Weight (kg)'] < len(r) else None
    dt = r[idx['Date']] if idx['Date'] < len(r) else None

    if member_no is None or kg is None or dt is None:
        skipped += 1
        continue

    try:
        kg_val = float(kg)
    except Exception:
        skipped += 1
        continue

    d, dto = as_date_and_time(dt)
    if not d or not dto:
        skipped += 1
        continue

    member_no_s = str(member_no).strip()
    member_name_s = str(member_name or '').strip()[:100]
    crop_s = str(crop or '').strip()[:30]
    bags = int(math.ceil(kg_val / 90.0))
    coll_no = f"{member_no_s}-{d.replace('-', '')}-{no}"[:30]

    payload = {
        'No': int(no),
        'Farmers_Number': member_no_s,
        'Farmers_Name': member_name_s,
        'Collections_Date': d,
        'Collection_Time': dto,
        'Collection_Number': coll_no,
        'Coffee_Type': crop_s[:20],
        'Kg_Collected': kg_val,
        'Gross': kg_val,
        'Tare': 0,
        'No_of_Bags': bags,
        'Factory': '',
        'Cancelled': False,
        'Paid': False,
        'Sent': False,
        'Updated': False,
        'ID_Number': '',
        'Delivered_By': member_name_s[:50],
        'Collect_Type': '',
        'Crop': crop_s,
        'Cumm': kg_val,
        'Can': '',
        'User': 'Philip',
        'Comments': 'Uploaded from paul use.xlsx'
    }

    body = json.dumps(payload).encode('utf-8')

    ok = False
    last_err = ''
    for attempt in range(1, 4):
        try:
            request_json(endpoint, method='POST', body=body, timeout=20)
            ok = True
            break
        except urllib.error.HTTPError as e:
            try:
                detail = e.read().decode('utf-8', 'ignore')
            except Exception:
                detail = str(e)
            last_err = f'HTTP {e.code} {detail[:200]}'
            # don't retry validation/key errors
            if e.code in (400, 401, 403, 404, 409, 422):
                break
            time.sleep(1.5)
        except Exception as e:
            last_err = str(e)[:200]
            time.sleep(1.5)

    if ok:
        created += 1
        no += 1
        if created % 25 == 0:
            print(f'progress created={created} failed={failed} skipped={skipped}')
    else:
        failed += 1
        if len(sample_errors) < 12:
            sample_errors.append(f'row {row_num} member {member_no_s}: {last_err}')
        # increment no to avoid repeated key errors if collision happened
        no += 1

print(f'UPLOAD_DONE created={created} failed={failed} skipped={skipped}')
if sample_errors:
    print('ERROR_SAMPLES_BEGIN')
    for e in sample_errors:
        print(e)
    print('ERROR_SAMPLES_END')

# verify count quickly
try:
    chk = request_json(endpoint + '?%24top=1')
    print('VERIFY_OK top_call_rows=', len(chk.get('value', [])))
except Exception as e:
    print('VERIFY_WARN', str(e)[:200])
