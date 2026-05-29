import json
import math
import base64
import urllib.request
import urllib.error
from datetime import datetime, date
import openpyxl

excel_path = r'C:\Users\mbugu\Downloads\paul use.xlsx'
endpoint = "http://test.trimline.co.ke:4548/BC240/ODataV4/Company('Rugi')/DailyCollections"
user = 'Philip'
pwd = 'Password@2030'

def auth_header():
    token = base64.b64encode(f"{user}:{pwd}".encode('ascii')).decode('ascii')
    return {
        'Authorization': f'Basic {token}',
        'Accept': 'application/json',
        'Content-Type': 'application/json'
    }

def get_start_no():
    url = endpoint + '?%24orderby=No%20desc&%24top=1'
    req = urllib.request.Request(url, headers=auth_header())
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode('utf-8', 'ignore'))
            rows = data.get('value', [])
            if rows:
                return int(rows[0].get('No', 0)) + 1
    except Exception:
        pass
    return 1

def as_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat(), v.strftime('%Y-%m-%dT%H:%M:%SZ')
    if isinstance(v, date):
        d = v.isoformat()
        return d, d + 'T00:00:00Z'
    s = str(v).strip()
    if not s:
        return None, None
    # try common parse
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date().isoformat(), dt.strftime('%Y-%m-%dT%H:%M:%SZ')
        except Exception:
            pass
    return None, None

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb[wb.sheetnames[0]]

headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}
required = ['Date', 'Member Number', 'Member Name', 'Crop Type', 'Net Weight (kg)']
missing = [h for h in required if h not in idx]
if missing:
    raise RuntimeError('Missing columns: ' + ', '.join(missing))

no = get_start_no()
created = 0
failed = 0
errors = []

for r in ws.iter_rows(min_row=2, values_only=True):
    if not r:
        continue
    member_no = r[idx['Member Number']]
    member_name = r[idx['Member Name']]
    crop = r[idx['Crop Type']]
    kg = r[idx['Net Weight (kg)']]
    dt = r[idx['Date']]

    if member_no is None or kg is None or dt is None:
        continue

    try:
        kg_val = float(kg)
    except Exception:
        continue

    d, dto = as_date(dt)
    if not d or not dto:
        continue

    member_no_s = str(member_no).strip()
    member_name_s = str(member_name or '').strip()[:100]
    crop_s = str(crop or '').strip()[:30]
    bags = int(math.ceil(kg_val / 90.0))
    coll_no = f"{member_no_s}-{d.replace('-', '')}-{no}"

    payload = {
        'No': int(no),
        'Farmers_Number': member_no_s,
        'Farmers_Name': member_name_s,
        'Collections_Date': d,
        'Collection_Time': dto,
        'Collection_Number': coll_no[:30],
        'Coffee_Type': crop_s[:20],
        'Kg_Collected': kg_val,
        'Gross': kg_val,
        'Tare': 0,
        'No_of_Bags': bags,
        'Factory': '',
        'Cancelled': False,
        'Paid': False,
        'Sent': False,
        'Updated': False,
        'ID_Number': '',
        'Delivered_By': member_name_s[:50],
        'Collect_Type': '',
        'Crop': crop_s,
        'Cumm': kg_val,
        'Can': '',
        'User': 'Philip',
        'Comments': 'Uploaded from paul use.xlsx'
    }

    body = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(endpoint, data=body, headers=auth_header(), method='POST')

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            _ = resp.read()
        created += 1
        no += 1
    except urllib.error.HTTPError as e:
        failed += 1
        try:
            emsg = e.read().decode('utf-8', 'ignore')
        except Exception:
            emsg = str(e)
        errors.append(f"No {no} member {member_no_s}: HTTP {e.code} {emsg[:400]}")
    except Exception as e:
        failed += 1
        errors.append(f"No {no} member {member_no_s}: {str(e)[:400]}")

print(f'Total attempted={created+failed} created={created} failed={failed}')
if errors:
    print('--- FIRST ERRORS ---')
    for line in errors[:10]:
        print(line)
